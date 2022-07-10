
import os
import re
# resuests モジュールをインポート
from ast import IsNot, Str
from contextlib import nullcontext
from operator import truediv

from bs4.builder import HTML
import openpyxl
from openpyxl.xml.constants import ACTIVEX, XLSX
from openpyxl.styles import PatternFill
import requests
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.common.utils import keys_to_typing
from selenium.webdriver.chrome.options import Options
import time

from selenium.webdriver.common.action_chains import ActionChains

import random
import configparser
import tkinter
import tkinter.messagebox as tkmsg



"""#フィールドを定義する

"""
config_ini = configparser.ConfigParser()
base = os.path.dirname(os.path.abspath(__file__))

#config.iniより設定をする
config_ini.read('config.ini', encoding='utf-8')
var2 = config_ini.get('DEFAULT', 'Driverpath')
var3 = config_ini.get('DEFAULT', 'InputWb')
var4 = config_ini.get('DEFAULT', 'Option')
var5 = config_ini.get('DEFAULT', 'WaitTime')



inputWb = openpyxl.load_workbook(var3)
inputWs = inputWb.worksheets[0]
reserchkaisu = inputWs.cell(row = 3, column = 11).value
max = inputWs.max_row
options = Options()

user_agent = ['Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/74.0.3729.169 Safari/537.36',
                  'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/72.0.3626.121 Safari/537.36',
                  'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/74.0.3729.157 Safari/537.36',
                  'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/60.0.3112.113 Safari/537.36'
                  ]
options = webdriver.ChromeOptions()
options.add_argument('--user-agent=' + user_agent[random.randrange(0, len(user_agent), 1)])
options.add_argument(var4)
driver = webdriver.Chromedriver = webdriver.Chrome(executable_path= var2 , chrome_options=options)
 
"""
関数　: ProductReserch       
機能　: 指定の値を検索してランキングを表示する。
        検索するもじれつが空の場合、次の行に移る
        含める。含めないによって　「PR」を含めるか含めないかを決定する。
        ページ番号によって、検索する範囲を指定する。

引数　:  num         検索行
        yoko        検索列
        startPage   検索ページ
        readCounter ランキングの順位
"""
def ProductReserch(num, yoko, startPage , ):
    #検索する行数を指定する
    for num in range(num , max + 1):
        
        #検索ワードを取得する
        reserchWord = inputWs.cell(row = num, column = yoko).value
        
        #カウント変数初期化
        readCounter = 0
        flag = 0
        if reserchWord is None:
        #検索ワードを記載していない時
            print("検索ワードを記載していません")


        else:
        #検索ワードを記載しているとき
            #指定したページ数繰り返し処理をする。
            for startPage in range(1 , reserchkaisu + 1 ):#num　+ 1
                #検索するURLを指定する。
                URL ="https://search.rakuten.co.jp/search/mall/" + str(reserchWord) + "/?p=" + str(startPage)
                time.sleep(int(var5))
                #URLを開く
                driver.get(URL)
                time.sleep(int(var5))
                #カレントページを取得する
                html =driver.page_source.encode('utf-8')
                #読み込む情報を解析する。
                info0 = BeautifulSoup(html, 'html.parser') 
                #指定のクラスを リストで取得する。
                taglist = info0.select("[class*='content title']")
                #含めないにした時のリスト
                removedTagList =[]
                #検索する正規表現を指定
                pattern = ".*PR.*"     

                #リストのアイテムを一つずつ実行
                for item in taglist:
                    #パターンにアイテムがマッチするか確認
                    res = re.match(pattern, str(item))
                    #マッチするときはなにもしない
                    if(res):
                        pass
                    #マッチしないとき
                    else:
                        #リストに追加する
                        removedTagList.append(item)


                #検索する商品名を指定する。
                productName = inputWs.cell(row = num, column = 1).value
                if "含める" == inputWs.cell(row = 2, column = 11).value:
                        #取得したリストを1つずつ表示
                        for tag in taglist:
                            #もじれつをインクリメント
                            readCounter = readCounter + 1
                            #リストにもじれつがあった場合
                            if str(productName) in str(tag):
                                #読み込み件数インクリメント
                                    inputWs.cell(row = num, column = yoko +1).value = readCounter
                                    #順位を表示する
                                    print(readCounter)
                                    flag =1
                            else:
                                pass    
                        if (flag == 1):
                        #見つかった場合は次の商品を検索する。
                            flag = 0
                            break
                #最後までなかった場合は圏外と入力
                inputWs.cell(row = num, column = yoko + 1).value = "圏外"

                
                        

                if "含めない" == inputWs.cell(row = 2, column = 11).value:
                        #取得したリストを1つずつ表示
                        for tag in removedTagList:
                            #もじれつをインクリメント
                            readCounter = readCounter + 1
                            #リストにもじれつがあった場合
                            if str(productName) in str(tag):
                                #読み込み件数インクリメント
                                    inputWs.cell(row = num, column = yoko +1).value = readCounter
                                    #順位を表示する
                                    print(readCounter)
                                    flag =1
                            else:
                                pass     
                        if (flag == 1):
                        #見つかった場合は次の商品を検索する。
                            flag = 0
                            break                     
                #最後までなかった場合は圏外と入力
                inputWs.cell(row = num, column = yoko + 1).value = "圏外"
    #Excelに保存
    inputWb.save('output_reserched.xlsx')


"""メインメソッドを定義
"""
if __name__ == "__main__":
    #C列の検索
    ProductReserch(3,3,1)
    #F列の検索
    ProductReserch(3,6,1)
    print("処理を終了します。")