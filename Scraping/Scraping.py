from bs4.builder import HTML
import openpyxl
from openpyxl.xml.constants import ACTIVEX, XLSX
from openpyxl.styles import PatternFill
import requests
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.common.utils import keys_to_typing
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.action_chains import ActionChains
import time



class Browser:
    def scrollByElemAndOffset(self, element, offset = 0):

        self.driver.execute_script("arguments[0].scrollIntoView();", element)

        if (offset != 0):
            script = "window.scrollTo(0, window.pageYOffset + " + str(offset) + ");"
            self.driver.execute_script(script)

#pathを設定 
driverPath ="C:\python\work\Scraping\chromedriver.exe"
browserPath = "https://news.yahoo.co.jp/articles/adccd19d8452f340ead2a525d57ecb2523c2eb97"







# ブラウザのオプションを格納する。
options = Options()
# Headlessモードを有効にする（コメントアウトするとブラウザが実際に立ち上がります）
#options.set_headless(True)


#Chromeを操作
driver = webdriver.Chromedriver = webdriver.Chrome(chrome_options=options,executable_path= driverPath)
driver.get(browserPath)



#スクロール
element = driver.find_element_by_id("articleCommentModule")
actions = ActionChains(driver)
actions.move_to_element(element)
actions.perform()

#在庫検索をクリック
#2秒待つ
time.sleep(2)
xpath= '/html/body/div[1]/div/main/div[1]/div/div[4]/div/aside/div[1]/div/div[3]/span/a/span'
elem_login_btn = driver.find_element_by_xpath(xpath)
elem_login_btn.click()

wb = openpyxl.Workbook()

#カレントページのURLを取得
cur_url = browser.current_url
time.sleep(1)
driver.get(cur_url)
html = driver.page_source.encode('utf-8')
time.sleep(1)
info01= BeautifulSoup(html, 'html.parser')
time.sleep(1)
info02= info01.find_all('div')
time.sleep(1)
#Excel貼り付け 
for news in info02:
    wb.active.cell(k, 1).value =  (news.getText()).replace(' ', '').replace('\n', '')
    k=k+1
    

#Excelに保存
wb.save('output_yahoonews.xlsx') 