
# resuests モジュールをインポート
from contextlib import nullcontext
from operator import truediv
import re
from bs4.builder import HTML
import openpyxl
from openpyxl.xml.constants import ACTIVEX, XLSX
from openpyxl.styles import PatternFill
import requests
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.common.utils import keys_to_typing
from selenium.webdriver.chrome.options import Options
import time
import datetime
from selenium.webdriver.common.action_chains import ActionChains
import chromedriver_binary


#変数を宣言する
#inputWb        : 取得するExcelブックを入力
#inputWs        : 取得するExcelシートを入力
#yoko       : 取得する値のセルを入力
#driverPath : クロムドライバーの格納先を入力  
#yoko       : 検索ワードの読み取り列を入力

inputWb = openpyxl.load_workbook('サンプル.xlsx')
inputWs = inputWb.worksheets[0]
outputWb = openpyxl.Workbook()
max = inputWs.max_row
driverPath = "chromedriver.exe"
yoko = 2
z = 1



for num in range(1 , max + 1):
    #変数を宣言する
    #resercWord : google検索ワード
    #options    : 画面を見せないときに使う機能

    reserchWord = inputWs.cell(row = num, column = yoko).value
    options = Options()


    outputWb.create_sheet(index = max , title= reserchWord )
    outputWb.active = outputWb.worksheets[ z ]
    z = z + 1 
    
     #google検索
    if len(inputWs.cell(row = num, column = yoko).value) != 0 :
        driver = webdriver.Chromedriver = webdriver.Chrome(chrome_options=options,executable_path= driverPath)
        driver.get('https://www.google.com/')  
        search = driver.find_element_by_name('q') 
        search.send_keys(reserchWord)            
        search.submit()

        #検索ワードがない場合
        if  reserchWord is None: 
           print("検索ワードがありません")
        #検索ワードがある場合
        else:
            # 1秒待つ
            time.sleep(1) 
            #変数を宣言する
            reserchedclik = inputWs.cell(row = num, column = 4).value

            #D列分だけクリックする
            for num2 in range(0 , reserchedclik ):
                try:
                    driver.execute_script("document.getESlementsByClassName('r21Kzd')["+ str(num2) +"].click()")
                    time.sleep(2)
                    driver.execute_script("document.getElementsByClassName('r21Kzd')["+ str(num2) +"].click()")
                except Exception:
                    time.sleep(1)

            #変数を宣言する
            #html      :　ページを取得する
            #info03    : 解析する
            #titleInfo : 質問内容のタイトル情報を取得する
            #naiyou   : 質問内容の内容情報を取得する

            html = driver.page_source.encode('utf-8')
            info03= BeautifulSoup(html, 'html.parser') 
            titleInfo= info03.find_all( attrs={ 'class': ['iDjcJe'] }) # OR検索
            naiyou= info03.find_all( attrs={ 'class': ['hgKElc', 'di3YZe'] })

            #Excel貼り付け 
            k = 1
            for news in titleInfo:
                outputWb.active.cell(k, 1).value =  (news.getText()).replace(' ', '').replace('\n', '')
                k = k + 1 
            c = 1
            for news in naiyou:
                outputWb.active.cell(c, 2).value =  (news.getText()).replace(' ', '').replace('\n', '')
                c = c + 1 

#Excelに保存
outputWb.save('output_reserched.xlsx')