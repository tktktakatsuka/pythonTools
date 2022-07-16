from ast import Not
from tkinter import messagebox
import pyautogui
import os
# resuests モジュールをインポート
from operator import truediv
from bs4.builder import HTML
from openpyxl.xml.constants import ACTIVEX, XLSX
from openpyxl.styles import PatternFill
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
import time
import random
import configparser
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import openpyxl
from selenium.webdriver.common.action_chains import ActionChains
import re
from tkinter import *
from tkinter import ttk
from tkinter import filedialog
import tkinter.font as f
'''


'''
# button1クリック時の処理
def button1_clicked():
    fTyp = [("","*")]
    iDir = os.path.abspath(os.path.dirname(__file__))
    filepath = filedialog.askopenfilename(filetypes = fTyp,initialdir = iDir)
    file1.set(filepath)


# フォルダ指定の関数
def dirdialog_clicked():
    iDir = os.path.abspath(os.path.dirname(__file__))
    iDirPath = filedialog.askdirectory(initialdir = iDir)
    entry1.set(iDirPath)

# ファイル指定の関数
def filedialog_clicked():
    fTyp = [("", "*")]
    iFile = os.path.abspath(os.path.dirname(__file__))
    iFilePath = filedialog.askopenfilename(filetype = fTyp, initialdir = iFile)
    entry2.set(iFilePath)

# 実行ボタン押下時の実行関数
def conductMain():
    text = ""

    dirPath = entry1.get()
    filePath = entry2.get()
    if dirPath:
        text += "フォルダパス：" + dirPath + "\n"
    if filePath:
        text += "ファイルパス：" + filePath

    if text:
        messagebox.showinfo("info", text)
    else:
        messagebox.showerror("error", "パスの指定がありません。")



'''
スクレイピングする関数
'''
def research(word,var3):
    ''''
    configsetting
    '''
    config_ini = configparser.ConfigParser()
    config_ini.read('C:\python\work\pythonTools\GoogleResercher\Google.ini', encoding='utf-8')
    var2 = config_ini.get('DEFAULT', 'Driverpath')
    #var3 = config_ini.get('DEFAULT', 'InputWb')
    



    '''
    excellsetting
    '''
    inputWb = openpyxl.load_workbook(var3)
    inputWs = inputWb[ '調査結果' ]
    num     = inputWs.max_row + 1
    yoko    = 0
    #reserchkaisu = inputWs.cell(row = 3, column = 11).value
    max = inputWs.max_row


    '''
    表題作成
    '''
    inputWs.cell(row = 1, column = yoko +1 ).value = '検索キーワード'
    #店名
    inputWs.cell(row = 1, column = yoko +2 ).value = '店舗名'
    #住所
    inputWs.cell(row = 1, column = yoko +3 ).value = '住所'
    #電話番号        
    inputWs.cell(row = 1, column = yoko +4 ).value =  '電話番号'
    #レビュー
    inputWs.cell(row = 1, column = yoko +5 ).value =  '口コミ評価'
    #レビュー件数
    inputWs.cell(row = 1, column = yoko +6 ).value =  '口コミ件数'


    """
    driversetting
    """
    config_ini = configparser.ConfigParser()
    base = os.path.dirname(os.path.abspath(__file__))
    path = os.getcwd()
    config_ini.read(path+'\\'+'Google.ini', encoding='utf-8')
    options = Options()
    user_agent = ['Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/74.0.3729.169 Safari/537.36',
                      'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/72.0.3626.121 Safari/537.36',
                      'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/74.0.3729.157 Safari/537.36',
                      'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/60.0.3112.113 Safari/537.36'
                      ]
    options = webdriver.ChromeOptions()
    options.add_argument('--user-agent=' + user_agent[random.randrange(0, len(user_agent), 1)])
    driver = webdriver.Chromedriver = webdriver.Chrome(executable_path= var2 , chrome_options=options)

    '''
    URLを開く
    '''
    URL ="https://www.google.co.jp/maps/?hl=ja"
    time.sleep(2)
    #URLを開く
    driver.get(URL)

    '''
    login処理
    '''
    # ページ上のすべての要素が読み込まれるまで待機（15秒でタイムアウト判定）
    WebDriverWait(driver, 15).until(EC.presence_of_all_elements_located)
    #メールアドレスを記入
    reserchWord     = driver.find_element_by_xpath('//*[@id="searchboxinput"]')
    reserchWord.clear()
    #word = '渋谷　焼肉'
    reserchWord.send_keys(word)

    time.sleep(3)
    xpath = '//*[@id="searchbox-searchbutton"]'
    elem = driver.find_element_by_xpath(xpath)
    elem.click()
    time.sleep(3)

    '''
    スクロール
    '''
    driver.maximize_window()
    pyautogui.moveTo(30, 200)
    for i in range(50):
        pyautogui.scroll(-100000)


        time.sleep(3)
    '''
    soupに解析させる
    '''
    #カレントページを取得する
    html =driver.page_source.encode('utf-8')
    #読み込む情報を解析する。
    soup = BeautifulSoup(html, 'html.parser') 
    #指定のクラスを リストで取得する。
    taglist = soup.select("[class*='Nv2PK THOPZb CpccDe']")




    '''
    情報抽出
    '''
    num     = inputWs.max_row + 1
    for item in taglist:

        storeURL       = item.find('a', class_='hfpxzc' )
        driver.get(str(storeURL.get('href')))
        time.sleep(1)
        




        '''
        soupに解析させる
        '''
        #カレントページを取得する
        html =driver.page_source.encode('utf-8')
        #読み込む情報を解析する。
        soup = BeautifulSoup(html, 'html.parser') 
        #指定のクラスを リストで取得する。
        taglist = soup.select("[class*='XltNde tTVLSc']")

        '''
        アイテム変数定義
        '''    
        for item in taglist:
            time.sleep(1)
            i = 1
            

            store       = item.find_all('div', class_='Io6YTe fontBodyMedium' )
            storeName   = item.find('h1', class_='DUwDvf fontHeadlineLarge' ).text
            address     = item.find('div', class_='AeaXub' ).text

            phoneNumberList =  soup.select("[class*='Io6YTe fontBodyMedium']")
            for phone in phoneNumberList:
                content = phone.text 
                pattern = '[0-9]*-[0-9]*-[0-9]*'
                result = re.search(pattern, str(content))
                if result: #none以外の場合
                    phoneNumber = str(phone.text)
                    

            Review       = driver.find_element_by_xpath('//*[@id="QA0Szd"]/div/div/div[1]/div[2]/div/div[1]/div/div/div[2]/div[1]/div[1]/div[2]/div/div[1]/div[2]')
            ReviewText = Review.text
            ReviewRating = item.find('button', class_='DkEaL' ).text

            '''
            アイテム記入場所定義
            '''
            #word
            inputWs.cell(row = num, column = yoko +1 ).value = word
            #店名
            inputWs.cell(row = num, column = yoko +2 ).value = storeName
            #住所
            inputWs.cell(row = num, column = yoko +3 ).value = address.replace(" ","")
            #電話番号        
            inputWs.cell(row = num, column = yoko +4 ).value =  str(phoneNumber)
            #レビュー
            inputWs.cell(row = num, column = yoko +5 ).value =  ReviewText
            #レビュー件数
            inputWs.cell(row = num, column = yoko +6 ).value =  ReviewRating  
            #次の行へとインクリメント
            print(i+ '件目' ,end=' ')
            print(storeName ,end=' ')
            print(address ,end=' ')
            print(phoneNumber ,end=' ')
            print(ReviewText ,end=' ')
            print(ReviewRating )
            num = num + 1
            i   = i   + 1




    #別名で保存
    inputWb.save(var3)
    #driver.quit()
    print("処理を終了します")


if __name__ == '__main__':
    root = Tk()
    root.title('goolgle店舗情報出力')
    root.geometry("600x300+600+300")
    

    #フレームの作成
    frame1 = ttk.Frame(root, padding=16)
    frame1.pack()
    #検索キーワード作成
    researchlabel = ttk.Label(frame1, text='検索キーワードを設定してください')
    researchlabel.pack(fill = 'x', padx=60, side = 'left')
    word = StringVar()
    entry1 = ttk.Entry(frame1, textvariable= word,width=50)
    entry1.pack(fill = 'x', padx=20, side = 'left')

    
    #参照するファイルパスの作成
    frame2 = ttk.Frame(root, padding=16)
    frame2.pack()
    label2 = ttk.Label(frame2, text='出力先のフォルダパスを指定してください（.xlsx）')
    label2.pack(fill = 'x', padx=20, side = 'left')
    output = StringVar()
    entry2 = ttk.Entry(frame2, textvariable=output,width=50)
    entry2.pack(fill = 'x', padx=20, side = 'left')


    #実行ボタンの作成
    frame3 = ttk.Frame(root, padding=16)
    frame3.pack()
    button = ttk.Button(frame3,text='OK',command=lambda: research(word.get(),output.get()))
    button.pack(fill = 'x', padx=20, side = 'left')


    # キャンセルボタンの設置
    button2 = ttk.Button(frame3, text=("閉じる"), command=quit)
    button2.pack(fill = 'x', padx=20, side = 'left')    


    # ウィンドウの表示開始
    root.mainloop()





#'''
#URLが押せた場合は押してずっとループしたい
#'''
#while driver.find_element_by_xpath(xpath)  is not None:
#
#    #次の検索結果に移動
#    xpath = '//*[@id="eY4Fjd"]/img'
#    elem = driver.find_element_by_xpath(xpath)
#    try:
#        elem.click()
#    except Exception:
#        break
#
#        
#    '''
#    スクロール
#    '''
#    time.sleep(1)
#    driver.maximize_window()
#    pyautogui.moveTo(30, 200)
#    for i in range(10):
#        pyautogui.scroll(-50000)
#    time.sleep(1)
#
#'''
#情報抽出
#'''
#for item in taglist:
#
#    storeURL       = item.find('a', class_='hfpxzc' )
#    driver.get(str(storeURL.get('href')))
#    time.sleep(1)
#
#
#    '''
#    soupに解析させる
#    '''
#    #カレントページを取得する
#    html =driver.page_source.encode('utf-8')
#    #読み込む情報を解析する。
#    soup = BeautifulSoup(html, 'html.parser') 
#    #指定のクラスを リストで取得する。
#    taglist = soup.select("[class*='XltNde tTVLSc']")
#
#    '''
#    アイテム変数定義
#    '''    
#    for item in taglist:
#        time.sleep(1)
#
#        store       = item.find_all('div', class_='Io6YTe fontBodyMedium' )
#        storeName   = item.find('h1', class_='DUwDvf fontHeadlineLarge' ).text
#        address     = item.find('div', class_='AeaXub' ).text
#
#        phoneNumberList =  soup.select("[class*='Io6YTe fontBodyMedium']")
#        for phone in phoneNumberList:
#            content = phone.text 
#            pattern = '[0-9]*-[0-9]*-[0-9]*'
#            result = re.search(pattern, str(content))
#            if result: #none以外の場合
#                phoneNumber = str(phone.text)
#                print(phoneNumber)
#        
#        Review       = driver.find_element_by_xpath('//*[@id="QA0Szd"]/div/div/div[1]/div[2]/div/div[1]/div/div/div[2]/div[1]/div[1]/div[2]/div/div[1]/div[2]')
#        ReviewText = Review.text
#        ReviewRating = item.find('button', class_='DkEaL' ).text
#
#        '''
#        アイテム記入場所定義
#        '''
#        #word
#        inputWs.cell(row = num, column = yoko +1 ).value = word
#        #店名
#        inputWs.cell(row = num, column = yoko +2 ).value = storeName
#        #住所
#        inputWs.cell(row = num, column = yoko +3 ).value = address
#        #電話番号        
#        inputWs.cell(row = num, column = yoko +4 ).value =  str(phoneNumber)
#        #レビュー
#        inputWs.cell(row = num, column = yoko +5 ).value =  ReviewText
#        #レビュー件数
#        inputWs.cell(row = num, column = yoko +6 ).value =  ReviewRating  
#        #次の行へとインクリメント
#        num = num + 1
#
#
#
#
##別名で保存
#inputWb.save(var3)
##driver.quit()
#print("処理を終了します")