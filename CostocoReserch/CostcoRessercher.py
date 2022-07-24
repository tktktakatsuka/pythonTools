
import os
# resuests モジュールをインポート
from operator import truediv
from bs4.builder import HTML
from openpyxl.xml.constants import ACTIVEX, XLSX
from openpyxl.styles import PatternFill
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.chrome.options import Options

import random
import configparser
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import openpyxl
import time
from selenium.webdriver.common.action_chains import ActionChains
from tkinter import *
from tkinter import ttk
from tkinter import messagebox
from openpyxl.styles import PatternFill


def CostcoResearcher(researchPage,outputFile,mailAddress,passwordText):

    
    '''
    excellsetting
    '''
    inputWb = openpyxl.Workbook()
    inputWs = inputWb.worksheets[0]
    num = 2
    yoko    = 0
    #reserchkaisu = inputWs.cell(row = 3, column = 11).value
    



    #%%
    """
    driversetting
    """
    config_ini = configparser.ConfigParser()
    config_ini.read('Costco.ini', encoding='utf-8')
    var2 = config_ini.get('DEFAULT', 'Driverpath')
    options = Options()
    user_agent = ['Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/74.0.3729.169 Safari/537.36',
                      'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/72.0.3626.121 Safari/537.36',
                      'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/74.0.3729.157 Safari/537.36',
                      'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/60.0.3112.113 Safari/537.36'
                      ]
    options = webdriver.ChromeOptions()
    options.add_argument('--user-agent=' + user_agent[random.randrange(0, len(user_agent), 1)])
    driver = webdriver.Chromedriver = webdriver.Chrome(executable_path= var2 , options=options)
    
    '''
    表題作成
    '''
    #注文番号取得
    inputWs.cell(row = 1, column = yoko +1 ).value = '注文番号'
    #商品名
    inputWs.cell(row = 1, column = yoko +2 ).value = '商品名'
    #ステータス
    inputWs.cell(row = 1, column = yoko +3 ).value = '配送状況'
    #数量
    inputWs.cell(row = 1, column = yoko +4 ).value = '数量'
    #注文日
    inputWs.cell(row = 1, column = yoko +5 ).value = '注文日'
    #住所        
    inputWs.cell(row = 1, column = yoko +7 ).value =  '住所'


    '''
    URLを開く
    '''
    URL ="https://www.costco.co.jp/login"

    #URLを開く
    driver.get(URL)


    '''
    login処理
    '''
    # ページ上のすべての要素が読み込まれるまで待機（15秒でタイムアウト判定）
    WebDriverWait(driver, 15).until(EC.presence_of_all_elements_located)
    #メールアドレスを記入
    mail     = driver.find_element_by_xpath('//*[@id="j_username"]')
    password = driver.find_element_by_xpath('//*[@id="j_password"]')
    mail.clear()
    password.clear()
    time.sleep(1)
    mail.send_keys(mailAddress)
    time.sleep(1)
    #print(password)
    password.send_keys(passwordText)
    xpath = '//*[@id="loginSubmit"]'
    elem = driver.find_element_by_xpath(xpath)
    elem.click()
    WebDriverWait(driver, 15).until(EC.presence_of_all_elements_located)  
    time.sleep(30)

    pathList = []
    pageNumber = 0
    #isTagList = True

    #while(isTagList):
    for i in range(int(researchPage) + 1):

        '''
        2つめのタブを開く
        '''
        # 新しいタブを作成する
        driver.execute_script("window.open()")
        # 新しいタブに切り替える
        driver.switch_to.window(driver.window_handles[1])
        # 新しいタブでURLアクセス
        URL ="https://www.costco.co.jp/my-account/orders?sort=byDate&page="+ str(pageNumber)
            #URLを開く
        driver.get(URL)
        pageNumber = pageNumber + 1


        '''
        soupに解析させる
        '''
        #カレントページを取得する
        html =driver.page_source.encode('utf-8')
        #読み込む情報を解析する。
        soup = BeautifulSoup(html, 'html.parser') 
        #指定のクラスを リストで取得する。
        taglist = soup.select("[class*='order-history__list-item']")



        '''
        情報抽出
        '''
        for item in taglist:
            '''
            アイテム変数定義
            '''
            pathNumber  = item.find('span', class_='order-id notranslate' ).text
            price       = item.find('div', class_='list-item__data ordr-history-price' ).text
            path = 'https://www.costco.co.jp/my-account/order/' + pathNumber.replace('\n','')
            pathList.append(path)

            #価格        
            inputWs.cell(row = num, column = yoko +6 ).value =  str(price)
            num = num +1


    #print(pathList)
    num =2

    #指定のクラスを リストで取得する。
    '''
    情報抽出
    '''


    for path in pathList:

        driver.get(str(path))
        WebDriverWait(driver, 15).until(EC.presence_of_all_elements_located)  
        '''
        soupに解析させる
        '''
        #カレントページを取得する
        html =driver.page_source.encode('utf-8')
        #読み込む情報を解析する。
        soup = BeautifulSoup(html, 'html.parser') 
        #指定のクラスを リストで取得する。
        productlist = soup.select("[class*='col-md-9 col-lg-10 no-space']")


        #注文番号取得                    
        orderNumber =  driver.find_element_by_xpath('//*[@id="selectItemsForm"]/div[1]/div[7]/div[1]/div/div[2]/div[1]/div[1]/div[2]/div/span')
        print(orderNumber.text)
        #商品名
        productName1 =  driver.find_element_by_xpath('//*[@id="selectItemsForm"]/div[1]/div[7]/div[3]/div/div[2]/div/div[3]/ul/li/div[1]/div[1]/div/div[2]/div/div[1]')
        print(productName1.text)
        productName2 =  driver.find_element_by_xpath('//*[@id="selectItemsForm"]/div[1]/div[7]/div[3]/div/div[2]/div/div[3]/ul/li/div[1]/div[1]/div/div[2]/div/div[2]')
        print(productName2.text)
        #status
        status = driver.find_element_by_xpath('//*[@id="selectItemsForm"]/div[1]/div[7]/div[1]/div/div[2]/div[2]/div/div/div/span')
        print(status.text)
        #数量
        quantity = driver.find_element_by_xpath('//*[@id="selectItemsForm"]/div[1]/div[7]/div[3]/div/div[2]/div/div[3]/ul/li/div[1]/div[1]/div/div[2]/div/div[6]/span')
        print(quantity.text)
        #注文日
        store  =driver.find_element_by_xpath('//*[@id="selectItemsForm"]/div[1]/div[7]/div[1]/div/div[2]/div[1]/div[1]/div[1]/div/span')
        print(store.text)
        #価格
        #price       =driver.find_element_by_xpath('//*[@id="selectItemsForm"]/div[1]/div[7]/div[4]/div/div/div/div/div[8]')
        #キャンセルの値段
        #price       =driver.find_element_by_xpath('//*[@id="selectItemsForm"]/div[1]/div[7]/div[4]/div/div/div/div/div[3]')
        #print(price.text)
        #住所　
        address       =driver.find_element_by_xpath('//*[@id="selectItemsForm"]/div[1]/div[7]/div[3]/div/div[2]/div/div[1]/div/div')
        print(address.text)
        '''
        アイテム記入場所定義
        '''
        #注文番号取得
        inputWs.cell(row = num, column = yoko +1 ).value = str(orderNumber.text)
        #商品名
        inputWs.cell(row = num, column = yoko +2 ).value = str(productName1.text + productName2.text)
        #ステータス
        inputWs.cell(row = num, column = yoko +3 ).value = str(status.text)
        #数量
        inputWs.cell(row = num, column = yoko +4 ).value = str(quantity.text)
        #注文日
        inputWs.cell(row = num, column = yoko +5 ).value = str(store.text)
        #住所        
        inputWs.cell(row = num, column = yoko +7 ).value =  str(address.text)
        
        num = num + 1
    inputWb.save(outputFile)
    # メッセージボックス（情報） 
    messagebox.showinfo('正常終了', '処理を終了します')



'''
mainメソッド
'''
if __name__ == '__main__':
    root = Tk()
    root.title('コストコ履歴調査')
    iconfile = 'sample.ico'
    root.iconbitmap(default=iconfile)
    root.geometry("600x350+500+300") 



    #検索ページ番号の指定
    frame1 = ttk.Frame(root, padding=16)
    label1 = ttk.Label(frame1, text='検索するページ数を指定してください')
    researchWord = StringVar()
    entry1 = ttk.Entry(frame1, textvariable=researchWord)
    frame1.pack()
    label1.pack()
    entry1.pack()

    #出力先ファイル名の指定
    frame2 = ttk.Frame(root, padding=16)
    label2 = ttk.Label(frame2, text='出力先のファイル名を指定してください'+ '\n' +'（ex:data.xlsx）')
    outputFile = StringVar()
    entry2 = ttk.Entry(frame2, textvariable=outputFile)
    frame2.pack()
    label2.pack()
    entry2.pack()

    #メールアドレス指定
    frame3 = ttk.Frame(root, padding=16)
    label3 = ttk.Label(frame3, text='mailAddress')
    mailAddress = StringVar()
    entry3 = ttk.Entry(frame3, textvariable=mailAddress)
    frame3.pack()
    label3.pack()
    entry3.pack()

    #Password指定
    frame4 = ttk.Frame(root, padding=16)
    label4 = ttk.Label(frame4, text='Password')
    password = StringVar()
    entry4 = ttk.Entry(frame4, textvariable=password)
    frame4.pack()
    label4.pack()
    entry4.pack()

    #ボタン設定
    button1 = ttk.Button(
        text='OK',
        command=lambda: CostcoResearcher(researchWord.get(),outputFile.get(),mailAddress.get(),password.get()))
    button1.pack()




    # ウィンドウの表示開始
    root.mainloop()