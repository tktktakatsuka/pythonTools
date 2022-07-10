
# resuests モジュールをインポート
from ast import IsNot
from contextlib import nullcontext
from operator import truediv
import re
from bs4.builder import HTML
import openpyxl
from openpyxl.xml.constants import ACTIVEX, XLSX
from openpyxl.styles import PatternFill
import requests
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.common.utils import keys_to_typing
from selenium.webdriver.chrome.options import Options
import time
import datetime
from selenium.webdriver.common.action_chains import ActionChains
import chromedriver_binary


"""変数を宣言する
#inputWb        : 取得するExcelブックを入力
#inputWs        : 取得するExcelシートを入力
#yoko       : 取得する値のセルを入力
#driverPath : クロムドライバーの格納先を入力  
#yoko       : 検索ワードの読み取り列を入力"""

inputWb = openpyxl.load_workbook('サンプル.xlsx')
inputWs = inputWb.worksheets[0]
outputWb = openpyxl.Workbook()
max = inputWs.max_row
driverPath = "C:\python\work\shopInfo/chromedriver.exe"
yoko = 2
z = 1



for num in range(1 , 1 + 1):#num　+ 1
    """#変数を宣言する
    #resercWord : google検索ワード
    #options    : 画面を見せないときに使う機能
    #URL最初のブラウザ画面"""

    reserchWord = inputWs.cell(row = num, column = yoko).value
    if  reserchWord is None:
        exit


    options = Options()
    URL ="https://www.shoppingmap.it/negozi/9846-harem.html"

    outputWb.create_sheet(index = max , title= reserchWord )
    outputWb.active = outputWb.worksheets[ z ]
    z = z + 1 
        #google検索
    if len(inputWs.cell(row = num, column = yoko).value) != 0:
        driver = webdriver.Chromedriver = webdriver.Chrome(chrome_options=options,executable_path= driverPath)
        driver.get(URL)  
        driver.maximize_window()
        time.sleep(5)
        #検索値を入力してクリック
        search = driver.find_element_by_xpath('/html/body/div[1]/div[1]/div/div[2]/div[1]/input') 
        search.send_keys(reserchWord)
        time.sleep(5)
        element = driver.find_element_by_xpath('/html/body/div[1]/div[1]/div/div[2]/div[2]/a[2]/span')             
        element.click()
        
        """変数を宣言する
        #html      :　ページを取得する
        #info03    : 解析する
        #titleInfo : 質問内容のタイトル情報を取得する
        # """

        html = driver.page_source.encode('utf-8')
        time.sleep(5)
        info03= BeautifulSoup(html, 'html.parser') 
        tags= info03.find_all (class_="social text-right" ) # OR検索

        # 　find_allはリスト形式で取得されるので、forで順番に取り出す。
        #　ポイントは取り出した情報からさらにタグ名を絞ってforで取り出すこと。
        
        for tag in tags:
            K = 1
            # 一つづつ取り出した〇〇〇クラスの中の"a"タグの情報を取得
            for a in tag.select("a"):
                # aタグの中の　title="タイトル名"　というようなtitle=のあとの情報を取得する
                title_name = a.get('title')
                print(title_name)
                # aタグの中の　href="リンク先のURL"　というようなhref=のあとの情報を取得する
                url = a.get('href')
                outputWb.active.cell(1, 1).value = url
                k = k + 1 
        
#Excelに保存
outputWb.save('output_reserched.xlsx')