
import re
from operator import truediv
import time
from bs4.builder import HTML
import openpyxl
from openpyxl.xml.constants import ACTIVEX, XLSX
from openpyxl.styles import PatternFill
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.common.utils import keys_to_typing
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.action_chains import ActionChains
import random
import configparser
import datetime

''''
configsetting
'''
config_ini = configparser.ConfigParser()
config_ini.read('C:\\python\\work\\pythonTool\\EthnicAsianShopResercher\\config.ini', encoding='utf-8')
var2 = config_ini.get('DEFAULT', 'Driverpath')
var3 = config_ini.get('DEFAULT', 'InputWb')
var4 = config_ini.get('DEFAULT', 'Option')

'''
excellsetting
'''
inputWb = openpyxl.load_workbook(var3,keep_vba=True)
inputWs = inputWb[ "調査結果" ]
num     = inputWs.max_row
yoko    = 0
#reserchkaisu = inputWs.cell(row = 3, column = 11).value
#max = inputWs.max_row

'''
chormedriversetrting
'''
options = Options()
user_agent = ['Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/74.0.3729.169 Safari/537.36',
                  'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/72.0.3626.121 Safari/537.36',
                  'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/74.0.3729.157 Safari/537.36',
                  'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/60.0.3112.113 Safari/537.36'
                  ]
options = webdriver.ChromeOptions()
options.add_argument('--user-agent=' + user_agent[random.randrange(0, len(user_agent), 1)])
options.add_argument(var4)
driver = webdriver.Chromedriver = webdriver.Chrome(executable_path= var2 , chrome_options=options)
 

'''
URL遷移
'''
URL ="https://www.asian-toybox.com/item_list.html"
#URLを開く


driver.get(URL)
time.sleep(5)
#カレントページを取得する
html =driver.page_source.encode('utf-8')

#読み込む情報を解析する。
soup = BeautifulSoup(html, 'html.parser') 
#指定のクラスを リストで取得する。
taglist = soup.select("[class*='item_list']")

del taglist[0]

'''
情報抽出
'''
for item in taglist:
    '''
    アイテム変数定義
    '''
    
    price   = item.find('p', class_='price')
    Product = item.find('p', class_='itemname')
    date   = item.find('p', class_='date')
    url   = item.find('a')

    if item.find('span', class_='tag_re') :
        stock = item.find('span', class_='tag_re')
    elif item.find('span', class_='tag_new'):
        stock = item.find('span', class_='tag_new')       
    else: 
        pass


    '''
    アイテム記入場所定義
    '''



    #入荷日の日付
    time_now = datetime.datetime.now()
    year = time_now.strftime("%Y")
    content = str(date.text)
    pattern    = '\d*/\d*'
    result = re.match(pattern, str(content))
    if result: #none以外の場合
        inputWs.cell(row = num, column = yoko +2).value =  year +"/" + str(result.group())
    
    #在庫状況
    content = stock
    regex = r'\".*\"'  # アスタリスクに囲まれている任意の文字
    mmatchStock = re.search(regex, str(content))  
    inputWs.cell(row = num, column = yoko +3).value = mmatchStock.group()

    #製品名
    inputWs.cell(row = num, column = yoko +4 ).value = str(Product.text)

    #価格  
    # ￥削除、コンマ削除
    sprice = str(price.text.replace('￥', '').replace(',', ''))
    inputWs.cell(row = num, column = yoko +5).value = sprice

    #URL情報
    inputWs.cell(row = num, column = yoko +6).value = str(url.get('href'))


    #モジュール実行日
    time_now = datetime.datetime.now()
    year = time_now.strftime("%Y")
    month = time_now.strftime("%m").lstrip("0")
    day = time_now.strftime("%d").lstrip("0")
    date_without_0 = year + "/" + month + "/" + day
    inputWs.cell(row = num, column = yoko +7).value = date_without_0

    #次の行へとインクリメント
    num = num + 1


'''
終了処理
'''
#上書きで保存
inputWb.save(var3)
driver.quit()
#ここまでoutputresercher



'''
置換作業
'''
#最終行
for row in inputWs.iter_rows():
    #
    for cell in row:
        ##入荷日
        if cell.col_idx == 2:
            pass

        #新規・再入荷
        if cell.col_idx == 3:
            new_text = cell.value.replace("\"tag_re\"", "再入荷").replace("\"tag_new\"", "新規")
            if new_text is None :
                pass
            else:
                cell.value = new_text

        #価格
        if cell.col_idx == 5:
            content = cell.value 
            pattern = '【.*】'
            result = re.search(pattern, str(content))
            if result: #none以外の場合
                cell.value = cell.value.replace(result.group(),"")


'''
9時間以上前だったら年を-1する
'''
yoko =2
max = inputWs.max_row
for tate in range(2 , max + 1):

    #入荷日を取得
    content = inputWs.cell(row = tate, column =  yoko ).value
    pattern = '/[0-9]*/'
    result = re.search(pattern, str(content))
    arrivalMonth = ""
    if result: #none以外の場合
        data = (result.group())
        arrivalMonth = data.replace("/","")
        

    #記入日を取得
    content = inputWs.cell(row = tate, column =  yoko  + 5 ).value
    pattern = '/[0-9]*/'
    result = re.search(pattern, str(content))
    executeMonth =""
    if result: #none以外の場合
        data = (result.group())
        executeMonth = data.replace("/","")
        

    if arrivalMonth:
        if executeMonth:
            subtraction =int(executeMonth) - int(arrivalMonth)
            
            if(subtraction <= -9):
            #9か月以上離れていた時
                #print (str(subtraction) + "時間差分があります。")
                #print(inputWs.cell(row = tate, column =  yoko ).value)

                content = inputWs.cell(row = tate, column =  yoko  + 5 ).value
                pattern = '20[0-9]*'
                result = re.search(pattern, str(content))
                iresult = int(result.group()) -1
                #print(iresult)
                inputWs.cell(row = tate, column =  yoko ).value =  inputWs.cell(row = tate, column =  yoko ).value.replace(str(result.group()) , str(iresult))

print('重複削除処理開始')
'''
重複行削除
'''
#先頭行からループ
for Q in range(inputWs.max_row + 1):
    print(str(Q) + '行目')
    if Q == 0:
        continue

    #セル値を変数へ格納
    arrivaldate = inputWs.cell(Q, 2).value
    product     = inputWs.cell(Q, 4).value
    
    #セル値の行番号を取得
    list_Num = Q

    #最終行から逆ループ
    for i in reversed(range(inputWs.max_row + 1)):
        if i == 0:
            break

        #セル値とlistが一緒だったら
        if str(inputWs.cell(i, 2).value)  in  str(arrivaldate) :
            if str(inputWs.cell(i, 4).value)  in  str(product) :

                #同じ行同士の比較はしない
                if i == Q:
                    continue
                else:
                    #行削除
                    inputWs.delete_rows(i)





#別名で保存
inputWb.save(var3)
print('処理を終了します。')







'''
取得するタグ
<div class="item_list re lady men" style="height: 297px;">
<a href="https://www.asian-toybox.com/item/8011" target="new">
<img src="https://image.rakuten.co.jp/asian-toybox/cabinet/banner/8011_330.jpg" alt="おうちウェアにもGOOD！スポーティーMIXパンツ"></a>
<p class="date">7/5</p>
<p class="tag">
<span class="tag_re"></span>
<span class="tag_lady"></span>
<span class="tag_men"></span>
</p>
<p class="itemname">【SPT】アジアンプリントポケット裾ドロストワイドパンツ</p>
<p class="price">￥4,752</p>
</div>
'''