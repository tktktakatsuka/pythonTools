
import re
from operator import truediv
from bs4.builder import HTML
import openpyxl
from openpyxl.xml.constants import ACTIVEX, XLSX
from openpyxl.styles import PatternFill
import datetime


'''
excellsetting
'''
inputWb = openpyxl.load_workbook('output_reserched.xlsm',keep_vba=True)
inputWs = inputWb[ "調査結果" ]


'''
置換作業
'''
##最終行
#for row in inputWs.iter_rows():
#    #
#    for cell in row:
#        #入荷日
#        if cell.col_idx == 2:
#            content = cell.value 
#            pattern    = '\d*/\d*'
#            result = re.match(pattern, str(content))
#            if result: #none以外の場合
#                today = datetime.date.today()
#                cell.value = str(today.strftime('%Y/')) + str(result.group())
#
#        #新規・再入荷
#        if cell.col_idx == 3:
#            new_text = cell.value.replace("\"tag_re\"", "再入荷").replace("\"tag_new\"", "新規")
#            cell.value = new_text
#
#        #価格
#        if cell.col_idx == 5:
#            content = cell.value 
#            pattern = '【.*】'
#            result = re.search(pattern, str(content))
#            if result: #none以外の場合
#                cell.value = cell.value.replace(result.group(),"")




#9時間以上前だったら年を-1する

yoko =2
max = inputWs.max_row
for tate in range(2 , max + 1):

    #入荷日を取得
    content = inputWs.cell(row = tate, column =  yoko ).value
    pattern = '/[0-9]*/'
    result = re.search(pattern, str(content))
    arrivalMonth = ""
    if result: #none以外の場合
        data = (result.group())
        arrivalMonth = data.replace("/","")
        

    #記入日を取得
    content = inputWs.cell(row = tate, column =  yoko  + 5 ).value
    pattern = '/[0-9]*/'
    result = re.search(pattern, str(content))
    executeMonth =""
    if result: #none以外の場合
        data = (result.group())
        executeMonth = data.replace("/","")
        

    if arrivalMonth:
        if executeMonth:
            subtraction =int(executeMonth) - int(arrivalMonth)
            
            if(subtraction <= -9):
            #9か月以上離れていた時
                #print (str(subtraction) + "時間差分があります。")
                #print(inputWs.cell(row = tate, column =  yoko ).value)

                content = inputWs.cell(row = tate, column =  yoko  + 5 ).value
                pattern = '20[0-9]*'
                result = re.search(pattern, str(content))
                iresult = int(result.group()) -1
                #print(iresult)
                inputWs.cell(row = tate, column =  yoko ).value =  inputWs.cell(row = tate, column =  yoko ).value.replace(str(result.group()) , str(iresult))


'''
重複行削除
'''
#先頭行からループ
for Q in range(inputWs.max_row + 1):
    if Q == 0:
        continue

    #セル値を変数へ格納
    arrivaldate = inputWs.cell(Q, 2).value
    product     = inputWs.cell(Q, 4).value
    
    

    #セル値の行番号を取得
    list_Num = Q

    #最終行から逆ループ
    for i in reversed(range(inputWs.max_row + 1)):
        if i == 0:
            break

        #セル値とlistが一緒だったら
        if str(inputWs.cell(i, 2).value)  in  str(arrivaldate) :
            if str(inputWs.cell(i, 4).value)  in  str(product) :

                #同じ行同士の比較はしない
                if i == Q:
                    continue
                else:
                    #行削除
                    inputWs.delete_rows(i)

list = list(range(max +1 ))
print(list)
for id in range(2 , max + 1):
    inputWs.cell(row = id, column =  1 ).value =  list[id - 1]





#別名で保存
inputWb.save('output_result.xlsm')



